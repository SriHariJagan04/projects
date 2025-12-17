import sys
import time
import vlc
import cv2
import warnings
import re
import pandas as pd
import csv
import os
import subprocess

from footer import Footer
from configPanel import ConfigPanel

from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal, QDate
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem,
    QFileDialog, QSlider, QFrame, QHeaderView, QAbstractItemView, QLabel,
    QSizePolicy, QSplitter, QMessageBox, QApplication
)

from PyQt5.QtGui import QFont, QImage, QPixmap

import random


import platform
import psutil


class BatchCompressor(QThread):
    def __init__(self, files_to_compress, output_dir):
        super().__init__()
        self.files_to_compress = files_to_compress
        self.output_dir = output_dir

    def run(self):
        for path in self.files_to_compress:
            base_name = os.path.splitext(os.path.basename(path))[0]
            compressed_filename = f"{base_name}_compressed_body.mp4"
            compressed_path = os.path.join(self.output_dir, compressed_filename)

            # Skip already compressed videos
            if path.endswith('_compressed_body.mp4') or os.path.exists(compressed_path):
                continue

            size_mb = os.path.getsize(path) / (1024 * 1024)
            if size_mb <= 30:
                continue

            compressor = VideoCompressor(path, compressed_path)
            compressor.run()


import subprocess
import os

class VideoCompressor(QThread):
    finished = pyqtSignal(bool, str)

    def __init__(self, input_path, output_path):
        super().__init__()
        self.input_path = input_path
        self.output_path = output_path

    def run(self):
        try:
            cmd = [
                "ffmpeg", "-i", self.input_path,
                "-vf", "scale=iw/2:ih/2",
                "-c:v", "libx264",
                "-preset", "ultrafast",
                "-crf", "35",
                "-r", "20",
                "-y",
                self.output_path
            ]

            # Hide CMD window on Windows
            startupinfo = None
            if os.name == "nt":
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW

            subprocess.run(
                cmd,
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                startupinfo=startupinfo
            )

            self.finished.emit(True, self.output_path)
        except subprocess.CalledProcessError:
            self.finished.emit(False, "")







class Body(QWidget):
    def __init__(self, config_data, header):
        super().__init__()

        self.header = header

        self.header.save_btn.clicked.connect(self.saveLogs)
        self.header.save_as_btn.clicked.connect(self.saveLogsAs)
        self.header.delete_btn.clicked.connect(self.clearLogs)
        

        self.config = config_data
        
        self.media_player = vlc.MediaPlayer()
        self.instance = vlc.Instance('--quiet', '--no-xlib')  # suppress log output
        self.media_player = self.instance.media_player_new()
        self.speed = 1.0
        self.last_speed_change = 0
        self.last_seek_time = 0

        self.pressed_keys = set()
        self.key_timer = QTimer(self)
        self.key_timer.setInterval(300)
        self.key_timer.timeout.connect(self.handle_key_hold)
        self.key_timer.start()

        # Video Frame
        self.video_frame = QFrame()
        self.video_frame.setStyleSheet("background-color: black;")
        self.video_frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Cross-platform video output
        QTimer.singleShot(100, self.set_video_output)

        # Seek Slider
        self.seek_slider = QSlider(Qt.Horizontal)
        self.seek_slider.setRange(0, 1000)
        self.seek_slider.setValue(0)
        self.seek_slider.sliderMoved.connect(self.set_position)

        self.timer = QTimer(self)
        self.timer.setInterval(30)
        self.timer.timeout.connect(self.update_ui)
        self.timer.start()

        self.time_label = QLabel("Time: 00:00")
        self.time_label.setAlignment(Qt.AlignCenter)


        self.upload_btn = QPushButton("📁 Upload Video")
        self.upload_btn.setStyleSheet("padding: 6px; font-weight: bold; background-color: #4CAF50; color: white; border-radius: 6px;")
        self.upload_btn.clicked.connect(self.upload_video)

        self.play_btn = QPushButton("▶ Play")
        self.play_btn.setCheckable(True)
        self.play_btn.setStyleSheet("padding: 6px; font-weight: bold; background-color: #2196F3; color: white; border-radius: 6px;")
        self.play_btn.clicked.connect(self.toggle_play)

        self.back_btn = QPushButton("⏪ Back")
        self.forward_btn = QPushButton("⏩ Forward")
        for btn in [self.back_btn, self.forward_btn]:
            btn.setStyleSheet("padding: 6px; font-weight: bold; background-color: #f0ad4e; color: white; border-radius: 6px;")
        self.back_btn.clicked.connect(self.seek_backward)
        self.forward_btn.clicked.connect(self.seek_forward)

        self.speed_label = QLabel("Speed: 1.00x")
        self.speed_label.setAlignment(Qt.AlignCenter)
        self.speed_label.setStyleSheet("font-weight: bold;")

        self.speed_down_btn = QPushButton("➖")
        self.speed_up_btn = QPushButton("➕")
        self.speed_down_btn.clicked.connect(self.decrease_speed)
        self.speed_up_btn.clicked.connect(self.increase_speed)
        for btn in [self.speed_down_btn, self.speed_up_btn]:
            btn.setFixedSize(40, 30)
            btn.setStyleSheet("font-size: 18px; font-weight: bold;")

        speed_layout = QHBoxLayout()
        speed_layout.addWidget(self.speed_down_btn)
        speed_layout.addWidget(self.speed_label)
        speed_layout.addWidget(self.speed_up_btn)

        control_layout = QHBoxLayout()
        control_layout.addWidget(self.upload_btn)
        control_layout.addWidget(self.play_btn)
        control_layout.addWidget(self.back_btn)
        control_layout.addWidget(self.forward_btn)
        control_layout.addLayout(speed_layout)

        video_layout = QVBoxLayout()
        video_layout.addWidget(self.video_frame)
        self.seek_time_layout = QHBoxLayout()
        self.seek_time_layout.setContentsMargins(0, 0, 0, 0)
        self.seek_time_layout.setContentsMargins(0, 0, 0, 0)
        self.seek_time_layout.setSpacing(10)
        self.seek_time_layout.addWidget(self.seek_slider)
        self.seek_time_layout.addWidget(self.time_label)
        video_layout.addLayout(self.seek_time_layout)
        video_layout.addLayout(control_layout)

        video_panel = QWidget()
        video_panel.setLayout(video_layout)
        video_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)


        self.table = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels(["TDate", "TTime", "WB", "DIR", "AXLES", "CLASS", "NAME","Avg Speed", ""])
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setColumnWidth(2, 50)
        self.table.setColumnWidth(3, 50)
        self.table.setColumnWidth(4, 60)
        self.table.setColumnWidth(5, 60)

        # Apply modern style
        self.table.setStyleSheet(""" 
            QTableWidget {
                border: 1px solid #4a4a4a;
                background-color: #ffffff;
                gridline-color: #dcdcdc;
                font-family: Segoe UI, sans-serif;
            }
            QHeaderView::section {
                background-color: #485e74;
                color: white;
                font-weight: bold;
                padding: 5px;
                border: none;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 4px
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)

        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(self.table)
        splitter.addWidget(video_panel)
        splitter.setStretchFactor(0, 2)  # Table
        splitter.setStretchFactor(1, 1)



       # Initialize Footer first
        self.footer = Footer(
            # media_player=self.media_player,
            play_toggle_func=self.toggle_play,
            seek_forward_func=self.seek_forward,
            seek_backward_func=self.seek_backward,
            increase_speed_func=self.increase_speed,
            decrease_speed_func=self.decrease_speed,
            header=None,
            add_row=self.add_row,  # ✅ Corrected
            body = self
        )


        self.video_date = None
        self.last_save_path = None
        self.last_save_format = None
        self.video_time = None
        self.final_video_time = None
        self.is_paused = False


        
        self.setFocusPolicy(Qt.StrongFocus)
        self.setFocus()

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(splitter)
        self.setLayout(main_layout)

        # Example avg speeds in km/h (mean, stddev)
        self.vehicle_avg_speed = {
            "h": (60, 5),   # HCV trucks
            "m": (50, 5),   # MAV
            "o": (40, 5),   # OSV
            "r": (25, 3),   # RICKS
            "c": (30, 4),   # CONS
            "a": (80, 10),  # Ambulance etc.
            "0": (35, 5),
            "z": (10, 2),   # Animal
            "v": (55, 7),   # Buses
            "1": (12, 2),   # Bicycle
            "2": (30, 5),   # 2W
            "3": (40, 6),   # 3W
            "4": (45, 6),
            "5": (60, 7),
            "6": (50, 5),
            "7": (45, 6),
            "8": (50, 6),
            "9": (55, 6),
        }
        
        self.diagnostics_logged = False
        self.load_backup()

        # OpenCV variables
        self.cv_video = None
        self.cv_frame_label = QLabel(self.video_frame)
        self.cv_frame_label.setAlignment(Qt.AlignCenter)
        self.cv_frame_label.setStyleSheet("background-color: black;")
        self.cv_frame_label.setVisible(False)
        self.cv_current_frame = 0




    def load_backup(self):
        path = "backup.csv"
        if not os.path.exists(path):
            return

        try:
            with open(path, newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader)  # Skip header
                for row in reader:
                    values = row[:8]
                    row_position = self.table.rowCount()
                    self.table.insertRow(row_position)
                    for col, value in enumerate(values):
                        item = QTableWidgetItem(value)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.table.setItem(row_position, col, item)

                    delete_btn = QPushButton("Delete")
                    delete_btn.clicked.connect(lambda _, btn=delete_btn: self.delete_row_from_button(btn))
                    self.table.setCellWidget(row_position, 8, delete_btn)
        except Exception as e:
            print("Error loading backup:", e)


    def delete_row_from_button(self, btn):
        for row in range(self.table.rowCount()):
            cell_widget = self.table.cellWidget(row, 8)  # Assuming column 8 has the button
            if cell_widget is btn:
                # ✅ Get values before removing
                values = [self.table.item(row, col).text() if self.table.item(row, col) else "" for col in range(8)]
                self.table.removeRow(row)
                self.remove_from_backup(values)
                break


        

    def set_video_output(self):
        win_id = int(self.video_frame.winId())
        if sys.platform.startswith("win"):
            self.media_player.set_hwnd(win_id)
        elif sys.platform.startswith("linux"):
            self.media_player.set_xwindow(win_id)
        elif sys.platform == "darwin":
            self.media_player.set_nsobject(win_id)



    def update_config(self, config_data):
        self.config = config_data
        video_path = config_data.get("video_path")
        if video_path:
            self.load_video(video_path)
            self.after_video_ready(video_path)
            self.video_time = None




    def load_video(self, video_path):
        if video_path:
            media = vlc.Media(video_path)
            media.add_option('--no-video-title-show')
            media.add_option('--file-caching=300')
            media.add_option('--avcodec-hw=none')
            media.add_option('--no-fastseek')
            self.media_player.set_media(media)
            self.media_player.pause()
            QTimer.singleShot(500, self.ensure_video_ready)

        else:
            warnings.warn("No video path provided. Please select a valid video file.")



    def ensure_video_ready(self):
        if not self.media_player.is_playing():
            self.media_player.play()
        self.play_btn.setChecked(True)
        self.play_btn.setText("⏸ Pause")





    def upload_video(self):
        self.last_save_format = None

        # Select a single video
        path, _ = QFileDialog.getOpenFileName(self, "Select Video", "", "Video Files (*.mp4 *.avi *.mkv)")
        if not path:
            return

        folder = os.path.dirname(path)
        base_name = os.path.splitext(os.path.basename(path))[0]
        compressed_filename = f"{base_name}_compressed_body.mp4"
        compressed_path = os.path.join(folder, compressed_filename)

        # ✅ Step 1: Use existing compressed file if available
        if os.path.exists(compressed_path):
            self.final_video_path = compressed_path
            self.after_video_ready(compressed_path)
        else:
            # ✅ Step 2: Compress selected file if needed
            size_mb = os.path.getsize(path) / (1024 * 1024)
            if size_mb <= 30:
                self.final_video_path = path
                self.after_video_ready(path)
            else:
                self.upload_btn.setText("Compressing selected video...")
                self.compressor_thread = VideoCompressor(path, compressed_path)
                self.compressor_thread.finished.connect(self.on_video_compression_done)
                self.compressor_thread.start()

        # ✅ Step 3: Background compress others (if not already compressed)
        all_videos = [
            os.path.join(folder, f)
            for f in os.listdir(folder)
            if f.lower().endswith(('.mp4', '.avi', '.mkv'))
        ]
        remaining_videos = [v for v in all_videos if v != path]

        if remaining_videos:
            self.batch_thread = BatchCompressor(remaining_videos, folder)
            self.batch_thread.start()


    def on_video_compression_done(self, success, path):
        if success:
            self.final_video_path = path
            self.upload_btn.setText("📁 Upload Video")
            self.after_video_ready(path)
        else:
            QMessageBox.critical(self, "Compression Failed", "Could not compress the video.")
            self.upload_btn.setText("📁 Upload Video")



    def after_video_ready(self, path):
        self.load_video(path)
        filename = os.path.basename(path)
        date, time = self.extract_timestamp_from_filename(filename)
        self.video_date = date
        self.video_time = time  # Make sure internal state is updated

        self.header.timestamps_input.setText(time)
        self.header.timestamps_input.repaint()
        QApplication.processEvents()  # Ensure UI updates

         # ✅ Add this line to initialize OpenCV for frame-accurate seek
        if self.cv_video is not None:
            self.cv_video.release()
        self.cv_video = cv2.VideoCapture(path)
        self.cv_current_frame = 0
     


    def toggle_play(self):
        if self.media_player.is_playing():
            self.media_player.pause()
            self.play_btn.setText("▶ Play")
            self.is_paused = True

            pos = self.media_player.get_time()
            self.time_label.setText(f"Time: {self.format_time(pos)}")

            if self.video_time is None:
                self.video_time = self.header.timestamps_input.text()

            h, m, s = map(int, self.video_time.split(":"))
            start_time_ms = (h * 3600 + m * 60 + s) * 1000
            total_time_ms = start_time_ms + pos
            self.final_video_time = self.format_time_hours(total_time_ms)

        else:
             # ⏸ Force sync after seek
            self.media_player.pause()
            QApplication.processEvents()

            # ⏳ Wait for VLC to apply last seek (max 200ms wait)
            QTimer.singleShot(100, self.resume_after_seek)


    def resume_after_seek(self):
        # Now get fresh time and play
        self.media_player.play()
        self.play_btn.setText("⏸ Pause")
        self.is_paused = False

        # ✅ Hide OpenCV frame overlay
        self.cv_frame_label.setVisible(False)







    def update_ui(self):
        buffer_value = 0  # default to 0

        if self.video_time is None or self.video_time.strip() == "":
            self.video_time = self.header.timestamps_input.text()

        if self.header.buffer_input is not None:
            try:
                buffer_value = int(self.header.buffer_input.text()) * 1000  # Convert seconds to ms
            except ValueError:
                buffer_value = 0  # Fallback if buffer input is invalid

        if self.media_player.get_length() > 0:
            pos = self.media_player.get_time()

            # Only update time_label if not paused
            if not self.is_paused:
                self.time_label.setText(f"Time: {self.format_time(pos)}")

            length = self.media_player.get_length()
            self.seek_slider.blockSignals(True)
            self.seek_slider.setValue(int((pos / length) * 1000))
            self.seek_slider.blockSignals(False)

            # Safely parse self.video_time
            try:
                h, m, s = map(int, self.video_time.strip().split(":"))
            except (ValueError, AttributeError):
                h, m, s = 0, 0, 0  # Default to 00:00:00
                self.video_time = "00:00:00"

            start_time_ms = (h * 3600 + m * 60 + s) * 1000
            total_time_ms = start_time_ms + pos + buffer_value

            # Wrap around 24 hours (86400 seconds = 86400000 ms)
            total_time_ms %= 86400000

            self.final_video_time = self.format_time_hours(total_time_ms)

            # Update the timestamp input in the UI
            self.header.timestamps_input.setText(self.final_video_time)





    def set_position(self, value):
        length = self.media_player.get_length()
        if length > 0:
            new_time = int((value / 1000) * length)
            self.media_player.set_time(new_time)

    def format_time(self, ms):
        seconds = int(ms / 1000)
        mins = seconds // 60
        secs = seconds % 60
        return f"{mins:02}:{secs:02}"
    



    def seek_backward(self):
        if self.cv_video is None or not self.cv_video.isOpened():
            print("OpenCV video not ready")
            return

        was_playing = self.media_player.is_playing()
        if was_playing:
            self.media_player.pause()
            QApplication.processEvents()

        now = time.time()
        if now - self.last_seek_time < 0.1:
            return
        self.last_seek_time = now

        vlc_time_ms = self.media_player.get_time()
        fps = self.cv_video.get(cv2.CAP_PROP_FPS)
        total_frames = self.cv_video.get(cv2.CAP_PROP_FRAME_COUNT)
        frame_estimate = int((vlc_time_ms / 1000) * fps)
        self.cv_current_frame = max(min(frame_estimate, total_frames - 1), 0)

        frames_to_seek = int(fps * 1)  # Seek 1 second back
        self.cv_current_frame = max(self.cv_current_frame - frames_to_seek, 0)
        self.cv_video.set(cv2.CAP_PROP_POS_FRAMES, self.cv_current_frame)
        ret, frame = self.cv_video.read()

        if ret:
            rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_image.shape
            bytes_per_line = ch * w
            qt_image = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qt_image)
            self.cv_frame_label.setPixmap(pixmap.scaled(
                self.video_frame.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation
            ))
            self.cv_frame_label.setVisible(True)
            self.time_label.setText(f"⏪ Frame: {self.cv_current_frame}")
        else:
            print("Failed to read frame")
            self.cv_current_frame = 0

        self.play_btn.setChecked(False)
        self.play_btn.setText("▶ Play")
        self.play_btn.setEnabled(False)
        QTimer.singleShot(100, lambda: self.play_btn.setEnabled(True))

        # Synchronize VLC to new position
        seek_time_ms = int((self.cv_current_frame / fps) * 1000)
        QApplication.processEvents()
        time.sleep(0.05)
        self.media_player.set_time(seek_time_ms)






    def seek_forward(self):
        step = 1000 
        pos = self.media_player.get_time()
        self.media_player.set_time(pos + step)

        if not self.media_player.is_playing():
            self.media_player.play()
            self.play_btn.setChecked(True)
            self.play_btn.setText("⏸ Pause")



    def increase_speed(self):
        now = time.time()
        if now - self.last_speed_change < 0.3:
            return
        
        self.last_speed_change = now
        current_rate = self.media_player.get_rate()

        if self.speed > current_rate:
            # Restore previous speed
            self.media_player.set_rate(self.speed)
            current_rate = self.speed  # 🔥 sync current_rate with what we just set
        else:
            # Normal increment
            current_rate = min(current_rate + 1.0, 16.0)
            self.media_player.set_rate(current_rate)

        self.speed = current_rate  # 🔥 always sync internal speed tracker
        self.speed_label.setText(f"Speed: {self.speed:.2f}x")


    def decrease_speed(self):
        now = time.time()
        if now - self.last_speed_change < 0.3:
            return
        self.last_speed_change = now
        self.speed = max(self.speed - 1.0, 1.0)
        self.media_player.set_rate(self.speed)
        self.speed_label.setText(f"Speed: {self.speed:.2f}x")



    def add_row(self, row_data, vehicle_code=None):

        path = self.config.get("video_path") or getattr(self, "final_video_path", None)
    
        if path:
            filename = os.path.basename(path)
            date, time = self.extract_timestamp_from_filename(filename)
        else:
            date, time = "", ""  # fallback if no valid video path

        # timestamp = self.media_player.get_time()
        video_Start_time = self.header.timestamps_input.text()
        final_time = None

        # Get buffer time in seconds and convert to milliseconds
        buffer_text = self.header.buffer_input.text()
        buffer_ms = int(buffer_text) * 1000 if buffer_text.isdigit() else 0

        # Convert hh:mm:ss to milliseconds
        h, m, s = map(int, video_Start_time.split(":"))
        start_time_ms = (h * 3600 + m * 60 + s) * 1000

        # Add video offset and buffer
        total_time_ms = start_time_ms + buffer_ms
        final_time = self.format_time_hours(total_time_ms)

        # Insert row into the table
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)

        newDate = self.config.get("date")

        if isinstance(newDate, QDate):
            formatted_date = newDate.toString("yyyy-MM-dd")
        else:
            formatted_date = QDate.currentDate().toString("yyyy-MM-dd")

        final_Date = date if date else formatted_date


        avg_speed = "N/A"
        if vehicle_code and vehicle_code in self.vehicle_avg_speed:
            mean_speed, stddev_speed = self.vehicle_avg_speed[vehicle_code]
            speed = random.gauss(mean_speed, stddev_speed)
            # Clamp speed to reasonable positive range
            speed = max(5, min(speed, 150))
            avg_speed = f"{speed:.1f}"  # Format to 1 decimal place

        # List of values to insert
        values = [
            final_Date,         # Column 0: Date
            final_time,         # Column 1: Final Time
            row_data[0],        # Column 2: WB
            row_data[1],        # Column 3: DIR
            row_data[2],        # Column 4: AXLES
            row_data[3],        # Column 5: CLASS
            row_data[4],        # Column 6: NAME
            avg_speed,
        ]

        # Set the values and center align them
        for col, value in enumerate(values):
            item = QTableWidgetItem(value)
            item.setFont(QFont("Segoe UI", 12))
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_position, col, item)

        # Add the "Delete" button to column 7
        delete_btn = QPushButton("Delete")
        delete_btn.clicked.connect(lambda _, r=row_position: self.delete_row(self.table.indexAt(delete_btn.pos()).row()))
        self.table.setCellWidget(row_position, 8, delete_btn)

        # Scroll to the new row, select it, and set focus
        self.table.scrollToItem(self.table.item(row_position, 0))
        self.table.selectRow(row_position)
        self.table.setFocus()

        self.append_to_backup(values) 



    def append_to_backup(self, values):
        backup_path = "backup.csv"
        longitude = self.config.get("longitude", "")
        latitude = self.config.get("latitude", "")
        try:
            is_new = not os.path.exists(backup_path)
            with open(backup_path, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                if is_new:
                    writer.writerow(["TDate", "TTime", "WB", "DIR", "AXLES", "CLASS", "NAME", "Avg Speed", "Longitude", "Latitude", "LANE"])
                writer.writerow(values + [longitude, latitude, ""])
        except Exception as e:
            print("Failed to append to backup:", e)




    def format_time_hours(self, ms):
        seconds = ms // 1000
        h = (seconds // 3600) % 24  # Wrap around 24-hour clock
        m = (seconds % 3600) // 60
        s = seconds % 60
        return f"{h:02}:{m:02}:{s:02}"



    

    def delete_row(self, row):
        self.table.removeRow(row)
        values = [self.table.item(row, col).text() if self.table.item(row, col) else "" for col in range(8)]
        self.remove_from_backup(values)


        row_count = self.table.rowCount()
        if row_count > 0:
            # Scroll to the new last row (or previous row if last was deleted)
            last_row = max(0, row_count - 1)
            self.table.scrollToItem(self.table.item(last_row, 0))
            # self.table.selectRow(last_row)


    def remove_from_backup(self, row_data):
        path = "backup.csv"
        if not os.path.exists(path):
            return

        try:
            with open(path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                lines = list(reader)

            headers = lines[0]
            updated_rows = [r for r in lines[1:] if r[:8] != row_data]

            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(updated_rows)

        except Exception as e:
            print("Failed to update backup file:", e)





    def keyPressEvent(self, event):
        key = event.key()
        self.pressed_keys.add(key)
        if key == Qt.Key_Space:
            self.toggle_play()

    def keyReleaseEvent(self, event):
        self.pressed_keys.discard(event.key())

    def handle_key_hold(self):
        if Qt.Key_B in self.pressed_keys:
            self.seek_backward()
        if Qt.Key_F in self.pressed_keys:
            self.seek_forward()
        if Qt.Key_Plus in self.pressed_keys or Qt.Key_Equal in self.pressed_keys:
            self.increase_speed()
        if Qt.Key_Minus in self.pressed_keys or Qt.Key_Underscore in self.pressed_keys:
            self.decrease_speed()


    def extract_timestamp_from_filename(self, filename):
        """
        Extract date and time from the filename. The format is expected to be:
        YYYYMMDDHHMMSS (e.g., 20240603120000)
        """
        match = re.match(r'(\d{8})(\d{6})', filename)  # Only one timestamp (YYYYMMDDHHMMSS)
        if match:
            start_date = match.group(1)  # First 8 digits: YYYYMMDD
            start_time = match.group(2)  # Next 6 digits: HHMMSS
            
            # Extract year, day, and month
            year = start_date[:4]
            day = start_date[4:6]
            month = start_date[6:]
            
            # Convert start date and time to the correct format
            formatted_date = f"{year}-{day}-{month}"  # YYYY-DD-MM
            formatted_time = f"{start_time[:2]}:{start_time[2:4]}:{start_time[4:]}"  # HH:MM:SS
            
            return formatted_date, formatted_time
        return None, None



        # Save to CSV
    def saveToCsv(self, filename, append=False):
        import csv
        rows = []
        longitude = self.config.get("longitude")
        latitude = self.config.get("latitude")

        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount() - 1):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            row_data.append(longitude)
            row_data.append(latitude)
            row_data.append("")  # Lane
            rows.append(row_data)

        try:
            with open(filename, mode='a' if append else 'w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                if not append:
                    writer.writerow(["TDate", "TTime", "WB", "DIR", "AXLES", "CLASS", "NAME", "Avg Speed", "Longitude", "Latitude", "LANE"])
                for row in rows:
                    writer.writerow(row)

            QMessageBox.information(self, "Success", "Logs saved successfully!", QMessageBox.Ok)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save logs to CSV: {str(e)}", QMessageBox.Ok)


    def saveToExcel(self, filename, append=False):
        from openpyxl import Workbook, load_workbook

        longitude = self.config.get("longitude")
        latitude = self.config.get("latitude")

        try:
            if append and os.path.exists(filename):
                wb = load_workbook(filename)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(["TDate", "TTime", "WB", "DIR", "AXLES", "CLASS", "NAME", "Avg Speed", "Longitude", "Latitude", "Lane"])

            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount() - 1):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                row_data.append(longitude)
                row_data.append(latitude)
                row_data.append("")  # Lane
                ws.append(row_data)

            wb.save(filename)
            QMessageBox.information(self, "Success", "Logs saved successfully!", QMessageBox.Ok)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save logs to Excel: {str(e)}", QMessageBox.Ok)




    def saveLogs(self):
        self.saveToCsv("backup.csv")


        
    # Save logs as (similar to saveLogs but with "Save As" functionality)
    def saveLogsAs(self):
        file_format = self.header.save_type.currentText().lower()
        filter_type = "CSV Files (*.csv)" if file_format == "csv" else "Excel Files (*.xlsx)"
        default_path = self.last_save_path or ""
        path, _ = QFileDialog.getSaveFileName(self, "Save As", default_path, filter_type)

        if path:
            self.last_save_path = path
            self.last_save_format = file_format

            if file_format == "csv":
                self.saveToCsv(path)
            else:
                self.saveToExcel(path)


    def clearLogs(self):
        self.table.setRowCount(0)
        try:
            open("backup.csv", "w").close()
        except Exception as e:
            print("Failed to clear backup file:", e)


        self.table.setRowCount(0)

