# save as kgpback_gui.py
import sys
import os
import hashlib
import pandas as pd
import re
import time
import subprocess
from pathlib import Path
from PyQt6.QtWidgets import (
    QApplication, QWidget, QGridLayout, QPushButton, QLabel, QFileDialog,
    QProgressBar, QMessageBox, QComboBox, QTableWidget, QTableWidgetItem,
    QTextEdit, QSizePolicy, QHeaderView
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ------------------------
# Utility functions
# ------------------------
def get_file_hash(file_path: str) -> str:
    h = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            h.update(chunk)
    return h.hexdigest()

def load_input_file(file_path: str) -> pd.DataFrame:
    ext = Path(file_path).suffix.lower()
    if ext == ".csv":
        return pd.read_csv(file_path, encoding="utf-8", engine="python")
    elif ext in [".xls", ".xlsx"]:
        return pd.read_excel(file_path, engine="openpyxl")
    raise ValueError(f"Unsupported file type: {ext}")

def extract_output_from_backout(file_path: str):
    # returns (surface, base, subgrade) strings or None
    surface = base = subgrade = None
    if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                if "Surface" in line:
                    m = re.search(r"=\s*([\d.]+)", line)
                    if m: surface = m.group(1)
                elif "Base" in line:
                    m = re.search(r"=\s*([\d.]+)", line)
                    if m: base = m.group(1)
                elif "Subgrade" in line:
                    m = re.search(r"=\s*([\d.]+)", line)
                    if m: subgrade = m.group(1)
    return surface, base, subgrade

def run_exe_and_wait(exe_path: str, input_values: list, timeout_seconds: int = 60):
    exe_dir = Path(exe_path).parent

    # 🧹 Remove old backout files
    for f in exe_dir.glob("backout*"):
        try:
            f.unlink()
        except Exception:
            pass

    # Prepare input text
    input_text = "\n".join(str(x) for x in input_values) + "\n\n"

    # 🪄 Hide console when running external EXE
    si = subprocess.STARTUPINFO()
    si.dwFlags |= subprocess.STARTF_USESHOWWINDOW

    # 🧠 CREATE_NO_WINDOW flag ensures no CMD popup appears
    proc = subprocess.Popen(
        [exe_path],
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        cwd=str(exe_dir),
        startupinfo=si,
        creationflags=subprocess.CREATE_NO_WINDOW
    )

    try:
        stdout, stderr = proc.communicate(input=input_text, timeout=timeout_seconds)
    except subprocess.TimeoutExpired:
        proc.kill()
        raise TimeoutError("Timeout waiting for EXE to finish.")

    # Wait for backout file
    start = time.time()
    out_file = None
    while time.time() - start < timeout_seconds:
        files = sorted(exe_dir.glob("backout*"), key=lambda p: p.stat().st_mtime)
        if files:
            cand = files[-1]
            if cand.exists() and cand.stat().st_size > 0:
                out_file = cand
                break
        time.sleep(0.8)

    if out_file is None:
        raise TimeoutError("Timeout waiting for backout file.")

    return extract_output_from_backout(str(out_file))

# ------------------------
# Worker Thread
# ------------------------
class WorkerThread(QThread):
    progress_signal = pyqtSignal(int)                   # percent int 0..100
    log_signal = pyqtSignal(str)                        # text log
    row_signal = pyqtSignal(int, list)                  # idx, full_result_row (list)
    finished_signal = pyqtSignal(pd.DataFrame)          # final dataframe
    error_signal = pyqtSignal(str)

    def __init__(self, input_file: str, exe_path: str, skip_rows=None, skip_cols=None, output_format="xlsx"):
        super().__init__()
        self.input_file = input_file
        self.exe_path = exe_path
        self.skip_rows = set(skip_rows or [])
        self.skip_cols = set(skip_cols or [])
        self.output_format = output_format
        self._stop = False

    def stop(self):
        self._stop = True

    def run(self):
        try:
            df = load_input_file(self.input_file)
            # normalize column names to str
            df.columns = [str(c) for c in df.columns]
            # expected columns (useful for assigning headers if transposed)
            expected_cols = [ "Load & TP", "N", "Spacing", "Deflection",
                             "Crust Thickness", "Poisson's ratio",
                             "BT Modulus", "GSB Modulus", "Subgrade Modulus"]
            

            orientation = "row"
            # if expected columns not present, assume input is column-wise -> transpose
            if not all(col in df.columns for col in expected_cols):
                df = df.transpose().reset_index(drop=True)
                # assign expected headers to leftmost columns if possible (preserve length)
                df.columns = expected_cols[:df.shape[1]]
                orientation = "column"

            columns_out = list(df.columns) + ["Surface (MPa)", "Base (MPa)", "Subgrade (MPa)"]
            output_file = Path(self.input_file).with_name(Path(self.input_file).stem + "_Results." + self.output_format)

            # Build set of existing input-rows in output (to skip). Compare only input columns (non-output columns)
            existing_keys = set()
            compare_cols = list(df.columns)  # input columns to compare
            if output_file.exists():
                try:
                    if self.output_format == "csv":
                        done_df = pd.read_csv(output_file, encoding="utf-8")
                    else:
                        done_df = pd.read_excel(output_file, engine="openpyxl")
                    # drop the 3 output columns if they exist in done_df
                    input_done_cols = [c for c in done_df.columns if c not in ["Surface (MPa)", "Base (MPa)", "Subgrade (MPa)"]]
                    # ensure columns align for comparison: we'll try to map by position if names differ
                    if list(input_done_cols) == compare_cols:
                        # straightforward
                        for _ , r in done_df[input_done_cols].iterrows():
                            key = tuple(str(v).strip() for v in r.values.tolist())
                            existing_keys.add(key)
                    else:
                        # fallback: compare by position using first len(compare_cols) columns of done_df
                        fallback_cols = list(done_df.columns[:len(compare_cols)])
                        for _ , r in done_df[fallback_cols].iterrows():
                            key = tuple(str(v).strip() for v in r.values.tolist())
                            existing_keys.add(key)
                except Exception:
                    existing_keys = set()

            total = len(df)
            last_percent = -1

            results = []   # we'll emit finished at the end

            for idx, row in df.iterrows():
                if self._stop:
                    break

                # prepare input values to send to EXE
                if orientation == "row":
                    # use column names, exclude skip_cols and "S.N."
                    cols_to_send = [
                        col for i, col in enumerate(df.columns)
                        if i not in self.skip_cols and col.strip().lower() != "s.n."
                    ]
                    input_values = [row[col] for col in cols_to_send]
                    # form compare key excluding S.N.
                    key = tuple(str(row[col]).strip() for col in compare_cols if col.strip().lower() != "s.n.")

                else:
                    # column-oriented: skip positions in skip_cols and "S.N." column
                    input_values = [
                        row.iloc[i] for i in range(len(row))
                        if i not in self.skip_cols and str(df.columns[i]).strip().lower() != "s.n."
                    ]
                    key = tuple(
                        str(row.iloc[i]).strip() for i in range(len(row))
                        if str(df.columns[i]).strip().lower() != "s.n."
                    )


                # skip if user manually selected this row to skip
                if idx in self.skip_rows:
                    self.log_signal.emit(f"Row {idx+1}: manually skipped")
                    # still append a placeholder row so final_df keeps same shape and ordering
                    results.append(list(row) + [None, None, None])
                    # update progress
                    pct = int((idx+1)/total*100)
                    if pct != last_percent:
                        last_percent = pct
                        self.progress_signal.emit(pct)
                    continue

                # skip if already present in results file
                if key in existing_keys:
                    self.log_signal.emit(f"Row {idx+1}: skipped (already present in results)")
                    results.append(list(row) + [None, None, None])  # will not overwrite existing outputs
                    pct = int((idx+1)/total*100)
                    if pct != last_percent:
                        last_percent = pct
                        self.progress_signal.emit(pct)
                    continue

                # run EXE and extract surface/base/subgrade
                try:
                    surface, base, subgrade = run_exe_and_wait(self.exe_path, input_values, timeout_seconds=60)
                    self.log_signal.emit(f"Row {idx+1}: Surface={surface}, Base={base}, Subgrade={subgrade}")
                except Exception as e:
                    surface = base = subgrade = None
                    self.log_signal.emit(f"Row {idx+1}: Error -> {e}")

                # final result row (input columns + outputs)
                result_row = list(row) + [surface, base, subgrade]
                results.append(result_row)

                # emit row-level result so GUI can update immediately
                self.row_signal.emit(idx, result_row)

                # append this single row to output file (dynamic save)
                df_row = pd.DataFrame([result_row], columns=columns_out)
                try:
                    if self.output_format == "csv":
                        header = not output_file.exists()
                        df_row.to_csv(output_file, index=False, mode="a", header=header, encoding="utf-8")
                    else:
                        if not output_file.exists():
                            df_row.to_excel(output_file, index=False, engine="openpyxl")
                        else:
                            book = load_workbook(output_file)
                            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                                writer.book = book
                                writer.sheets = {ws.title: ws for ws in book.worksheets}
                                startrow = writer.sheets['Sheet1'].max_row
                                df_row.to_excel(writer, index=False, header=False, startrow=startrow)
                                writer.save()
                    # add to existing_keys to avoid reprocessing same row in same run / duplicates
                    existing_keys.add(key)
                except Exception as e:
                    self.log_signal.emit(f"Row {idx+1}: Failed to append to output file -> {e}")

                # update progress intermittently
                pct = int((idx+1)/total*100)
                if pct != last_percent:
                    last_percent = pct
                    self.progress_signal.emit(pct)

            # after loop: final DF (keeping same ordering as input, with outputs where computed)
            final_df = pd.DataFrame(results, columns=columns_out)
            # deduplicate final_df (just in case) keeping first occurrence
            try:
                deduped = final_df.drop_duplicates(keep="first")
                final_df = deduped
            except Exception:
                pass

            self.finished_signal.emit(final_df)

        except Exception as e:
            self.error_signal.emit(str(e))


# ------------------------
# Main GUI
# ------------------------
class KGPBackGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("KGPBACK Processor")
        self.setMinimumSize(1100, 700)
        self.setAcceptDrops(True)

        self.setStyleSheet("""
            QWidget { background-color: #1e1e2f; color: #ffffff; font-family: Arial; font-size: 13px; }
            QPushButton { background-color: #2f2f4f; padding: 8px; border-radius: 6px; }
            QPushButton:hover { background-color: #45456e; }
            QLabel { font-size: 13px; }
            QProgressBar { border: 1px solid #ffffff; border-radius: 6px; text-align: center; }
            QProgressBar::chunk { background-color: #00bfff; }
            QTextEdit { background-color: #2b2b3b; color: #aaffdd; }
            QTableWidget { background-color: #222233; color: #ffffff; }
        """)

        layout = QGridLayout()
        layout.setSpacing(12)

        # Controls
        self.label_input = QLabel("Input File: Not selected")
        self.btn_input = QPushButton("Select Input")
        self.btn_input.clicked.connect(self.select_input_file)

        self.label_exe = QLabel("EXE File: Not selected")
        self.btn_exe = QPushButton("Select EXE")
        self.btn_exe.clicked.connect(self.select_exe_file)

        self.combo_format = QComboBox()
        self.combo_format.addItems(["csv", "xlsx"])
        self.combo_format.setToolTip("Select output format for results file")

        self.btn_start = QPushButton("Start Processing")
        self.btn_start.clicked.connect(self.start_processing)

        self.progress = QProgressBar()
        self.progress.setValue(0)

        self.table = QTableWidget()
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)

        self.log_window = QTextEdit()
        self.log_window.setReadOnly(True)
        self.log_window.setMaximumHeight(220)

        self.status_label = QLabel("")   # used to show "already processed" status
        self.status_label.setStyleSheet("color: lightgreen; font-weight: bold;")

        # place widgets
        layout.addWidget(self.btn_input, 0, 0)
        layout.addWidget(self.label_input, 0, 1)
        layout.addWidget(self.btn_exe, 1, 0)
        layout.addWidget(self.label_exe, 1, 1)
        layout.addWidget(QLabel("Output format:"), 2, 0)
        layout.addWidget(self.combo_format, 2, 1)
        layout.addWidget(self.btn_start, 3, 0, 1, 2)
        layout.addWidget(self.progress, 4, 0, 1, 2)
        layout.addWidget(QLabel("Input Data Preview (select rows/cols to skip):"), 5, 0, 1, 2)
        layout.addWidget(self.table, 6, 0, 1, 2)
        layout.addWidget(QLabel("Processing Log:"), 7, 0, 1, 2)
        layout.addWidget(self.log_window, 8, 0, 1, 2)
        layout.addWidget(self.status_label, 9, 0, 1, 2)

        self.setLayout(layout)

        self.input_file = None
        self.exe_path = None
        self.worker = None
        self.current_df = None
        self.existing_processed_rows = set()  # indices already present in output file

    # drag & drop
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
    def dropEvent(self, event):
        for url in event.mimeData().urls():
            f = url.toLocalFile()
            if Path(f).suffix.lower() in [".csv", ".xls", ".xlsx"]:
                self.input_file = f
                self.label_input.setText(f"Input File: {Path(f).name}")
                self.load_table_preview()
                break

    def select_input_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "Select input CSV/Excel", "", "Excel Files (*.xlsx *.xls);;CSV Files (*.csv)")
        if file:
            self.input_file = file
            self.label_input.setText(f"Input File: {Path(file).name}")
            self.load_table_preview()

    def select_exe_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "Select EXE file", "", "Executable Files (*.exe)")
        if file:
            self.exe_path = file
            self.label_exe.setText(f"EXE File: {Path(file).name}")

    def load_table_preview(self):
        try:
            if not self.input_file:
                return
            df = load_input_file(self.input_file)
            df.columns = [str(col) for col in df.columns]
            self.current_df = df.copy()

            # if transposed/column oriented preview is possible (we keep preview as-is; worker will transpose if necessary)
            rows = df.shape[0]
            cols = df.shape[1]
            self.table.setRowCount(rows)
            self.table.setColumnCount(cols)
            self.table.setHorizontalHeaderLabels([str(c) for c in df.columns])
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

            for i in range(rows):
                for j, col in enumerate(df.columns):
                    self.table.setItem(i, j, QTableWidgetItem(str(df.iat[i, j])))

        except Exception as e:
            self.log_window.append(f"Failed to load preview: {e}")

    def start_processing(self):
        if not self.input_file or not self.exe_path:
            QMessageBox.warning(self, "Warning", "Please select both input and EXE files!")
            return

        out_fmt = self.combo_format.currentText()
        out_path = Path(self.input_file).with_name(Path(self.input_file).stem + "_Results." + out_fmt)
        hash_file = Path(self.input_file).with_name(Path(self.input_file).stem + "_hash.txt")
        count_file = Path(self.input_file).with_name(Path(self.input_file).stem + "_count.txt")

        try:
            current_hash = get_file_hash(self.input_file)
            input_df = pd.read_excel(self.input_file) if self.input_file.lower().endswith((".xlsx", ".xls")) else pd.read_csv(self.input_file)
            input_count = len(input_df)

            # ✅ Check if previous hash and counts exist
            if hash_file.exists() and count_file.exists() and out_path.exists():
                with open(hash_file, "r") as fh:
                    old_hash = fh.read().strip()
                with open(count_file, "r") as fc:
                    old_count = int(fc.read().strip())

                # ✅ Only skip if hash matches AND all entries processed
                if old_hash == current_hash:
                    try:
                        if out_fmt.lower() == "csv":
                            out_df = pd.read_csv(out_path)
                        elif out_fmt.lower() in ("xlsx", "xls"):
                            out_df = pd.read_excel(out_path)
                        else:
                            out_df = None

                        if out_df is not None and len(out_df) >= input_count:
                            QMessageBox.information(self, "Skipped", "All entries already processed earlier. Skipping execution.")
                            self.status_label.setText("✅ All entries processed; skipped.")
                            return
                    except Exception as e:
                        self.log_window.append(f"Warning: Failed to verify output completeness -> {e}")
            
            # 🧾 Save new hash and count before starting
            with open(hash_file, "w") as fh:
                fh.write(current_hash)
            with open(count_file, "w") as fc:
                fc.write(str(input_count))

        except Exception as e:
            self.log_window.append(f"Warning: failed to compute/save hash/count -> {e}")

        # collect skip rows/cols from user selection
        skip_rows = [r.row() for r in self.table.selectionModel().selectedRows()]
        skip_cols = [c.column() for c in self.table.selectionModel().selectedColumns()]

        self.btn_start.setText("Processing...")
        self.btn_start.setEnabled(False)
        self.log_window.clear()
        self.progress.setValue(0)

        # start worker thread
        self.worker = WorkerThread(
            self.input_file,
            self.exe_path,
            skip_rows=skip_rows,
            skip_cols=skip_cols,
            output_format=out_fmt
        )
        self.worker.progress_signal.connect(self.progress.setValue)
        self.worker.log_signal.connect(lambda txt: self.log_window.append(txt))
        self.worker.row_signal.connect(self._on_row_processed)
        self.worker.finished_signal.connect(self._on_finished)
        self.worker.error_signal.connect(self._on_error)
        self.worker.start()


    def _on_row_processed(self, idx: int, result_row: list):
        """Update preview table live for the processed row and ensure output columns are visible."""
        try:
            # make sure preview table has space for output columns
            in_cols = self.current_df.shape[1] if self.current_df is not None else 0
            out_cols = 3
            total_cols_needed = in_cols + out_cols
            if self.table.columnCount() < total_cols_needed:
                # extend headers
                headers = [self.table.horizontalHeaderItem(i).text() if self.table.horizontalHeaderItem(i) else f"Col{i}" for i in range(in_cols)]
                headers += ["Surface (MPa)", "Base (MPa)", "Subgrade (MPa)"]
                self.table.setColumnCount(total_cols_needed)
                self.table.setHorizontalHeaderLabels(headers)
            # set cells for this row (only within visible preview size)
            for j, val in enumerate(result_row):
                # only update if within current preview rows/cols - but extend if necessary
                if idx < self.table.rowCount():
                    # for safety convert None to empty string
                    s = "" if val is None else str(val)
                    # ensure cell exists
                    if self.table.item(idx, j) is None:
                        self.table.setItem(idx, j, QTableWidgetItem(s))
                    else:
                        self.table.item(idx, j).setText(s)
            # mark processed row background to indicate done
            for j in range(self.table.columnCount()):
                item = self.table.item(idx, j)
                if item:
                    item.setBackground(Qt.GlobalColor.darkGreen)
        except Exception as e:
            self.log_window.append(f"UI update error: {e}")

    def _on_finished(self, final_df: pd.DataFrame):
        # when worker finished: ensure final output saved nicely and formatted
        try:
            out_fmt = self.combo_format.currentText()
            out_path = Path(self.input_file).with_name(Path(self.input_file).stem + "_Results." + out_fmt)
            # if worker already appended rows per row, final_df may only hold processed rows; we still write back a de-duplicated complete file:
            try:
                if out_fmt == "csv":
                    out_df = pd.read_csv(out_path, encoding="utf-8") if out_path.exists() else final_df
                else:
                    out_df = pd.read_excel(out_path, engine="openpyxl") if out_path.exists() else final_df
            except Exception:
                out_df = final_df
            # dedupe & save
            before = len(out_df)
            out_df = out_df.drop_duplicates(keep="first")
            after = len(out_df)
            if out_fmt == "csv":
                out_df.to_csv(out_path, index=False, encoding="utf-8")
            else:
                out_df.to_excel(out_path, index=False, engine="openpyxl")

            # apply Excel formatting if xlsx
            if out_fmt == "xlsx":
                try:
                    wb = load_workbook(out_path)
                    ws = wb.active
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    center = Alignment(horizontal="center", vertical="center")
                    for col in range(1, ws.max_column+1):
                        cell = ws.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center
                    # auto width
                    for col_cells in ws.columns:
                        length = max(len(str(c.value or "")) for c in col_cells)
                        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length + 2
                    wb.save(out_path)
                except Exception as e:
                    self.log_window.append(f"Warning: failed to apply excel formatting -> {e}")

            self.log_window.append(f"Finished. Results saved: {out_path} (removed {before-after} duplicates)")
            self.status_label.setText("✅ Processing completed.")
            QMessageBox.information(self, "Done", f"Processing complete.\nResults saved to:\n{out_path}")
        except Exception as e:
            self.log_window.append(f"Finalize error: {e}")
        finally:
            self.btn_start.setText("Start Processing")
            self.btn_start.setEnabled(True)
            self.progress.setValue(100)

    def _on_error(self, msg: str):
        QMessageBox.critical(self, "Error", msg)
        self.btn_start.setText("Start Processing")
        self.btn_start.setEnabled(True)
        self.log_window.append(f"ERROR: {msg}")

# ------------------------
# Run App
# ------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = KGPBackGUI()
    w.show()
    sys.exit(app.exec())
